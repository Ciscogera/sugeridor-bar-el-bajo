import os
import difflib
import re
import pandas as pd
import openpyxl
import streamlit as st
import io
import random
import string
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

# --- CONFIGURACIÓN DE LA PÁGINA WEB ---
st.set_page_config(page_title="Pedidos El Bajo", page_icon="", layout="centered")

# --- ⚠️ CONFIGURACIÓN DE REDIRECCIÓN OAUTH ---
# Recuerda cambiar esta URL por la definitiva cuando lo subas a Streamlit Cloud
REDIRECT_URI = "https://sugeridor-bar-el-bajo.streamlit.app/"

# --- BASE DE DATOS DE COLUMNAS DE PROVEEDORES ---
CONFIG_PROVEEDORES = {
    "Desa": {"col_nombre": 1, "col_pedido": 2, "fila_inicio": 2},
    "CCU": {"col_nombre": 1, "col_pedido": 4, "fila_inicio": 2},
    "ANDINA": {"col_nombre": 1, "col_pedido": 3, "fila_inicio": 2},
    "Tost": {"col_nombre": 1, "col_pedido": 3, "fila_inicio": 2},
    "ATF (TH Tonicas)": {"col_nombre": 1, "col_pedido": 2, "fila_inicio": 2},
    "Tubinger": {"col_nombre": 1, "col_pedido": 3, "fila_inicio": 2},
    "Vinoteca": {"col_nombre": 1, "col_pedido": 2, "fila_inicio": 2},
    "Cerros de Chena": {"col_nombre": 1, "col_pedido": 2, "fila_inicio": 2},
    "Tamango": {"col_nombre": 1, "col_pedido": 2, "fila_inicio": 2},
    "Kombuchacha": {"col_nombre": 1, "col_pedido": 2, "fila_inicio": 2},
    "ByMaria": {"col_nombre": 1, "col_pedido": 2, "fila_inicio": 2},
    "TeGusta": {"col_nombre": 1, "col_pedido": 2, "fila_inicio": 2},
    "Limache": {"col_nombre": 1, "col_pedido": 2, "fila_inicio": 2},
    "Segafredo": {"col_nombre": 1, "col_pedido": 2, "fila_inicio": 4},
}

# --- AUXILIAR GENERADOR DE CLAVES (ANTI-RESET) ---
def generar_state_pkce(length=64):
    chars = string.ascii_letters + string.digits
    return "".join(random.choice(chars) for _ in range(length))

# --- CONTROL DE SESIÓN GENERAL ---
if "credentials" not in st.session_state:
    st.session_state.credentials = None
if "auth_procesada" not in st.session_state:
    st.session_state.auth_procesada = False
if "etapa" not in st.session_state:
    st.session_state.etapa = "upload"
if "ambiguedades" not in st.session_state:
    st.session_state.ambiguedades = {}
if "cache_decisiones" not in st.session_state:
    st.session_state.cache_decisiones = {}
if "inventario_db" not in st.session_state:
    st.session_state.inventario_db = {}
if "pedidos_bytes" not in st.session_state:
    st.session_state.pedidos_bytes = None
if "excel_final" not in st.session_state:
    st.session_state.excel_final = None

# --- CAPTURA DE RETORNO GOOGLE (OAUTH HANDSHAKE) ---
if st.session_state.credentials is None and not st.session_state.auth_procesada:
    query_params = st.query_params
    if "code" in query_params and "state" in query_params:
        try:
            code = query_params["code"]
            state_retornado = query_params["state"]
            
            if "google_secrets" in st.secrets:
                client_config = {"web": dict(st.secrets["google_secrets"])}
                flow = Flow.from_client_config(
                    client_config,
                    scopes=['https://www.googleapis.com/auth/drive.readonly'],
                    redirect_uri=REDIRECT_URI,
                    code_verifier=state_retornado
                )
            elif os.path.exists('client_secrets.json'):
                flow = Flow.from_client_secrets_file(
                    'client_secrets.json',
                    scopes=['https://www.googleapis.com/auth/drive.readonly'],
                    redirect_uri=REDIRECT_URI,
                    code_verifier=state_retornado
                )
            else:
                flow = None
                
            if flow is not None:
                flow.fetch_token(code=code)
                st.session_state.credentials = flow.credentials
                st.session_state.auth_procesada = True
                st.session_state.etapa = "upload"
                st.query_params.clear()
                st.rerun()
                
        except Exception as e:
            st.error(f"⚠️ Error en intercambio de llaves de Google: {e}")
            st.session_state.auth_procesada = False
            st.session_state.etapa = "upload"

# --- FUNCIONES AUXILIARES DE GOOGLE DRIVE API ---
def listar_archivos_excel():
    try:
        service = build('drive', 'v3', credentials=st.session_state.credentials)
        query = "mimeType = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' and trashed = false"
        resultados = service.files().list(q=query, fields="files(id, name)", pageSize=50).execute()
        archivos = resultados.get('files', [])
        return {f['name']: f['id'] for f in archivos}
    except Exception as e:
        st.error(f"Error al conectar con Drive: {e}")
        return {}

def descargar_archivo_desde_drive(file_id):
    service = build('drive', 'v3', credentials=st.session_state.credentials)
    request = service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
    fh.seek(0)
    return fh

# --- LÓGICA DE PROCESAMIENTO DE STOCK ---
def cargar_inventario_real(file_io):
    df = pd.read_excel(file_io, sheet_name="Inventario General", skiprows=4)
    df = df.dropna(subset=["Nombre Producto"])
    df = df[df["Nombre Producto"].str.strip() != ""]
    df["Par Stock"] = pd.to_numeric(df["Par Stock"], errors="coerce").fillna(0)
    df["Bodega"] = pd.to_numeric(df["Bodega"], errors="coerce").fillna(0)
    df["Barra"] = pd.to_numeric(df["Barra"], errors="coerce").fillna(0)
    df["Total Actual"] = df["Bodega"] + df["Barra"]
    
    # Agrupamos por producto para evitar pérdida de registros duplicados
    df_agrupado = df.groupby(df["Nombre Producto"].str.strip()).agg({
        "Par Stock": "sum",
        "Total Actual": "sum"
    }).reset_index()
    
    return {str(r["Nombre Producto"]): {"par": r["Par Stock"], "actual": r["Total Actual"]} for _, r in df_agrupado.iterrows()}

def normalizar_texto_producto(texto):
    """Limpia el texto, remueve muletillas de inventario y estandariza unidades."""
    t = str(texto).lower().strip()
    t = re.sub(r"[()\-.,_']", " ", t)
    # Remueve palabras comodín de inventario que los proveedores no usan
    t = re.sub(r"\b(normal|tradicional|comun|estandar)\b", " ", t)
    t = re.sub(r"\b(750\s*ml|750\s*cc|750)\b", "750cc", t)
    t = re.sub(r"\b(1\s*l|1\s*litro|litro|1000\s*ml|1000\s*cc|1000)\b", "litro", t)
    t = re.sub(r"\b(330\s*ml|330\s*cc)\b", "330cc", t)
    t = re.sub(r"\b(355\s*ml|355\s*cc)\b", "355cc", t)
    t = re.sub(r"\b(500\s*ml|500\s*cc)\b", "500cc", t)
    return re.sub(r"\s+", " ", t).strip()

def encontrar_coincidencia_inteligente(nombre_prov, lista_inv):
    n_prov_raw = nombre_prov.lower().strip()
    n_prov_norm = normalizar_texto_producto(nombre_prov)
    
    # 1. Coincidencia exacta
    for item in lista_inv:
        if n_prov_raw == item.lower().strip() or n_prov_norm == normalizar_texto_producto(item):
            return item, "PERFECTO"
            
    # 2. Sub-texto con desempate por especificidad de nombre
    exactas = []
    for item in lista_inv:
        item_norm = normalizar_texto_producto(item)
        if item_norm and (n_prov_norm in item_norm or item_norm in n_prov_norm):
            exactas.append(item)
            
    if len(exactas) == 1: 
        return exactas[0], "PERFECTO"
    if len(exactas) > 1: 
        # Priorizar el texto de mayor longitud para diferenciar variantes (ej: "Flora Adora" vs "Hendricks")
        max_len = max(len(normalizar_texto_producto(x)) for x in exactas)
        mas_especificos = [x for x in exactas if len(normalizar_texto_producto(x)) == max_len]
        if len(mas_especificos) == 1:
            return mas_especificos[0], "PERFECTO"
        return exactas, "DUPLICADO"
        
    # 3. Fuzzy Match sobre cadenas normalizadas (Case-Insensitive)
    mapa_norm = {normalizar_texto_producto(x): x for x in lista_inv}
    mejores_norm = difflib.get_close_matches(n_prov_norm, list(mapa_norm.keys()), n=3, cutoff=0.3)
    
    if not mejores_norm: 
        return list(lista_inv), "BAJA_CERTEZA"
    
    mejores_orig = [mapa_norm[m] for m in mejores_norm]
    ratio = difflib.SequenceMatcher(None, n_prov_norm, mejores_norm[0]).ratio()
    
    if ratio < 0.85:
        return mejores_orig, "BAJA_CERTEZA"
    return mejores_orig[0], "ALTA_CERTEZA"

def ejecutar_calculo_matematico():
    wb = openpyxl.load_workbook(io.BytesIO(st.session_state.pedidos_bytes))
    inventario = st.session_state.inventario_db
    
    for sheet_name in wb.sheetnames:
        if sheet_name not in CONFIG_PROVEEDORES: continue
        ws = wb[sheet_name]
        conf = CONFIG_PROVEEDORES[sheet_name]
        
        for row in range(conf["fila_inicio"], ws.max_row + 1):
            cell_val = ws.cell(row=row, column=conf["col_nombre"]).value
            if not cell_val: continue
            n_prov = str(cell_val).strip()
            if n_prov.lower() in ["productos", "producto", "total", "rut:", "detalle de producto"]: continue
            
            # Usamos la clave única por fila
            clave_unica = f"{sheet_name}_f{row}_{n_prov}"
            item_elegido = st.session_state.cache_decisiones.get(clave_unica)
            
            if item_elegido and item_elegido in inventario:
                datos = inventario[item_elegido]
                cantidad = max(0, datos["par"] - datos["actual"])
                ws.cell(row=row, column=conf["col_pedido"]).value = cantidad if cantidad > 0 else ""
            else:
                ws.cell(row=row, column=conf["col_pedido"]).value = ""
                
    buffer = io.BytesIO()
    wb.save(buffer)
    st.session_state.excel_final = buffer.getvalue()

# --- UNIFICACIÓN DE ANÁLISIS Y ESCÁNER ---
def procesar_escaner_ambiguedades(io_inv, io_ped):
    st.session_state.inventario_db = cargar_inventario_real(io_inv)
    st.session_state.pedidos_bytes = io_ped.getvalue() if hasattr(io_ped, "getvalue") else io_ped.read()
    
    nombres_inv = list(st.session_state.inventario_db.keys())
    wb_scan = openpyxl.load_workbook(io.BytesIO(st.session_state.pedidos_bytes))
    ambiguedades_encontradas = {}
    
    for sheet_name in wb_scan.sheetnames:
        if sheet_name not in CONFIG_PROVEEDORES: continue
        ws = wb_scan[sheet_name]
        conf = CONFIG_PROVEEDORES[sheet_name]
        
        for row in range(conf["fila_inicio"], ws.max_row + 1):
            cell_val = ws.cell(row=row, column=conf["col_nombre"]).value
            if not cell_val: continue
            n_prov = str(cell_val).strip()
            if n_prov.lower() in ["productos", "producto", "total", "rut:", "detalle de producto"]: continue
            
            clave_unica = f"{sheet_name}_f{row}_{n_prov}"
            res, tipo_match = encontrar_coincidencia_inteligente(n_prov, nombres_inv)
            
            if tipo_match in ["ALTA_CERTEZA", "PERFECTO"]:
                st.session_state.cache_decisiones[clave_unica] = res
            else:
                # Captura DUPLICADO, BAJA_CERTEZA y NINGUNO para resolver en la interfaz
                ambiguedades_encontradas[clave_unica] = {
                    "nombre_prov": n_prov,
                    "hoja": sheet_name,
                    "candidatos": res if isinstance(res, list) else [res],
                    "tipo": tipo_match
                }
                
    st.session_state.ambiguedades = ambiguedades_encontradas
    if ambiguedades_encontradas:
        st.session_state.etapa = "resolver"
    else:
        ejecutar_calculo_matematico()
        st.session_state.etapa = "descargar"
    st.rerun()

# --- INTERFAZ DE USUARIO ---
st.title("Pedidos Automáticos - El Bajo")

# --- ETAPA 1: OBTENCIÓN DE ARCHIVOS ---
if st.session_state.etapa == "upload":
    st.subheader("1. Selección del Origen de Planillas")
    tab_drive, tab_local = st.tabs(["🔗 Google Drive (Nube)", "📂 Archivos Locales (PC/Móvil)"])
    
    with tab_drive:
        if st.session_state.credentials is None:
            st.write("Conéctate de forma segura a Google Drive para listar tus planillas de stock.")
            mi_state_secreto = generar_state_pkce()
            
            if "google_secrets" in st.secrets:
                client_config = {"web": dict(st.secrets["google_secrets"])}
                flow = Flow.from_client_config(
                    client_config,
                    scopes=['https://www.googleapis.com/auth/drive.readonly'],
                    redirect_uri=REDIRECT_URI,
                    code_verifier=mi_state_secreto
                )
            elif os.path.exists('client_secrets.json'):
                flow = Flow.from_client_secrets_file(
                    'client_secrets.json',
                    scopes=['https://www.googleapis.com/auth/drive.readonly'],
                    redirect_uri=REDIRECT_URI,
                    code_verifier=mi_state_secreto
                )
            else:
                flow = None
                
            if flow is not None:
                auth_url, _ = flow.authorization_url(prompt='select_account', state=mi_state_secreto)
                st.link_button("🔑 CONECTAR CON GOOGLE DRIVE", auth_url, use_container_width=True)
            else:
                st.error("Falta el archivo 'client_secrets.json' o la configuración en Secrets de Streamlit.")
        else:
            st.success("🟢 Cuenta vinculada exitosamente.")
            with st.spinner("Leyendo archivos de tu Google Drive..."):
                diccionario_archivos = listar_archivos_excel()
                
            if diccionario_archivos:
                opciones_excel = ["-- Seleccionar un archivo --"] + list(diccionario_archivos.keys())
                archivo_inv_name = st.selectbox("Elija el Inventario Diario Digitalizado:", options=opciones_excel, key="drive_inv")
                archivo_ped_name = st.selectbox("Elija el Maestro de Pedidos (Plantilla):", options=opciones_excel, key="drive_ped")
                
                if archivo_inv_name != "-- Seleccionar un archivo --" and archivo_ped_name != "-- Seleccionar un archivo --":
                    if st.button("🔍 ANALIZAR PLANILLAS DE DRIVE", use_container_width=True):
                        try:
                            id_inv = diccionario_archivos[archivo_inv_name]
                            id_ped = diccionario_archivos[archivo_ped_name]
                            with st.spinner("Descargando y escaneando datos del Bar..."):
                                io_inv = descargar_archivo_desde_drive(id_inv)
                                io_ped = descargar_archivo_desde_drive(id_ped)
                                procesar_escaner_ambiguedades(io_inv, io_ped)
                        except Exception as e:
                            st.error(f"Fallo en lectura de celdas de Drive: {e}")
            else:
                st.warning("No se encontraron archivos Excel (.xlsx) en la raíz de tu Google Drive.")
                if st.button("🔄 Intentar reconectar cuenta"):
                    st.session_state.credentials = None
                    st.session_state.auth_procesada = False
                    st.rerun()

    with tab_local:
        st.write("Sube tus archivos directamente desde tu computador o celular sin vincular ninguna cuenta.")
        archivo_inv_local = st.file_uploader("Sube el Inventario Diario Digitalizado (.xlsx)", type=["xlsx"], key="local_inv")
        archivo_ped_local = st.file_uploader("Sube el Maestro de Pedidos original (.xlsx)", type=["xlsx"], key="local_ped")
        
        if archivo_inv_local and archivo_ped_local:
            if st.button("🔍 ANALIZAR PLANILLAS LOCALES", use_container_width=True):
                try:
                    io_inv = io.BytesIO(archivo_inv_local.read())
                    io_ped = io.BytesIO(archivo_ped_local.read())
                    procesar_escaner_ambiguedades(io_inv, io_ped)
                except Exception as e:
                    st.error(f"Fallo en procesamiento local: {e}")

# --- ETAPA 2: RESOLVER AMBIGÜEDADES ---
elif st.session_state.etapa == "resolver":
    st.subheader("⚠️ Validación de Nombres")
    st.info(f"Faltan confirmar {len(st.session_state.ambiguedades)} productos:")
    
    with st.form("formulario_resolucion_drive"):
        nuevas_decisiones = {}
        bloque_items = list(st.session_state.ambiguedades.items())[:12]
        
        for clave_unica, info in bloque_items:
            n_prov = info["nombre_prov"]
            hoja = info["hoja"]
            tipo_texto = "Duplicado" if info["tipo"] == "DUPLICADO" else "Certeza Baja"
            options_select = ["[ No pedir ]"] + info["candidatos"]
            nuevas_decisiones[clave_unica] = st.selectbox(
                label=f"📋 '{n_prov}' (Hoja: {hoja} | {tipo_texto})", 
                options=options_select, 
                key=clave_unica
            )
            st.markdown("---")
            
        if st.form_submit_button("💾 GUARDAR ASOCIACIONES Y CONTINUAR", use_container_width=True):
            for clave_unica, eleccion in nuevas_decisiones.items():
                st.session_state.cache_decisiones[clave_unica] = None if eleccion == "[ No pedir ]" else eleccion
                del st.session_state.ambiguedades[clave_unica]
            
            if not st.session_state.ambiguedades:
                ejecutar_calculo_matematico()
                st.session_state.etapa = "descargar"
            st.rerun()

# --- ETAPA 3: DESCARGA FINAL ---
elif st.session_state.etapa == "descargar":
    st.subheader("🎯 ¡Sugerencia de Pedidos Completada!")
    st.success("La inyección matemática se completó de manera exitosa.")
    
    st.download_button(
        label="📥 DESCARGAR PEDIDO BARRA PROCESADO",
        data=st.session_state.excel_final,
        file_name="Pedido_Barra_Drive_Final.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
    
    if st.button("🔄 Procesar Nuevas Planillas", use_container_width=True):
        credenciales_actuales = st.session_state.credentials
        auth_actual = st.session_state.auth_procesada
        st.session_state.clear()
        st.session_state.credentials = credenciales_actuales
        st.session_state.auth_procesada = auth_actual
        st.session_state.etapa = "upload"
        st.rerun()